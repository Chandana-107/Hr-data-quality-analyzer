import pandas as pd
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def autofit(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ''
                max_length = max(max_length, len(v))
            except Exception:
                pass
        adjusted = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted


def _write_df_to_sheet(writer, df: pd.DataFrame, sheet_name: str):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    # freeze header
    ws.freeze_panes = 'A2'
    # header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # autofilter
    ws.auto_filter.ref = ws.dimensions
    autofit(ws)


def generate_report(output_path: Path, df: pd.DataFrame, results: dict, meta: dict):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Executive summary
        summary = {
            'Report generated at': [meta.get('generated_at')],
            'Source file': [meta.get('source_file')],
            'Total records': [meta.get('total_records')],
            'Overall score': [meta.get('overall_score')],
            'Classification': [meta.get('classification')]
        }
        summary_df = pd.DataFrame(summary)
        summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
        ws = writer.sheets['Executive Summary']
        for cell in ws[1]:
            cell.font = Font(bold=True)
        autofit(ws)

        # Completeness sheet
        comp = results.get('completeness', {})
        fields = comp.get('fields', {})
        rows = []
        for f, v in fields.items():
            rows.append({'Field': f, 'Total Records': v['total'], 'Populated': v['populated'], 'Missing': v['missing'], 'Completeness %': v['completeness_pct']})
        comp_df = pd.DataFrame(rows)
        _write_df_to_sheet(writer, comp_df, 'Completeness')

        # Validation Results
        val = results.get('validity', {})
        val_rows = pd.DataFrame(val.get('row_issues', []))
        if val_rows.empty:
            val_rows = pd.DataFrame(columns=['employee_id','field','value','rule','category','severity','message'])
        _write_df_to_sheet(writer, val_rows, 'Validation Results')

        # Duplicates
        dup = results.get('uniqueness', {})
        dup_rows = pd.DataFrame(dup.get('row_issues', []))
        if dup_rows.empty:
            dup_rows = pd.DataFrame(columns=['field','value','employee_id','rule','severity','message'])
        # transform duplicates summary
        _write_df_to_sheet(writer, dup_rows, 'Duplicates')

        # Consistency Issues
        cons = results.get('consistency', {})
        cons_df = pd.DataFrame(cons.get('row_issues', []))
        if cons_df.empty:
            cons_df = pd.DataFrame(columns=['employee_id','field','actual','expected','rule','severity','message'])
        _write_df_to_sheet(writer, cons_df, 'Consistency Issues')

        # Anomalies
        an = results.get('anomalies', {})
        an_df = pd.DataFrame(an.get('row_issues', []))
        if an_df.empty:
            an_df = pd.DataFrame(columns=['employee_id','anomaly','value','severity','message'])
        _write_df_to_sheet(writer, an_df, 'Anomalies')

        # Row-level issues combined (for convenience)
        combined = []
        for k in ('completeness','uniqueness','validity','consistency','anomalies'):
            r = results.get(k, {})
            df_r = r.get('row_issues')
            if isinstance(df_r, pd.DataFrame) and not df_r.empty:
                df_copy = df_r.copy()
                df_copy['category'] = k
                combined.append(df_copy)
        if combined:
            combined_df = pd.concat(combined, ignore_index=True, sort=False).fillna('')
        else:
            combined_df = pd.DataFrame(columns=['employee_id','field','value','rule','category','severity','message'])
        _write_df_to_sheet(writer, combined_df, 'Row Issues')

        # optionally include a small sample of source data
        try:
            sample = df.head(500)
            _write_df_to_sheet(writer, sample.reset_index(drop=True), 'Source Sample')
        except Exception:
            pass
